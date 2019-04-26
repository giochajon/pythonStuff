
from copy import copy


from openpyxl import load_workbook, Workbook


def copySheet(sourceSheet1, sourceSheet2, newSheet):
    for row in sourceSheet1.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                                    value=cell.value)
    last_row = cell.row - 1
    for row in enumerate(sourceSheet2.iter_rows(min_row=2), start=2):
        for cell in row[1]:
            new_row = last_row+cell.row
            newCell = newSheet.cell(row=new_row, column=cell.col_idx,
                                    value=cell.value)


filepaths = ['./invoice.xlsx', './invoice2.xlsx']

filepath = [load_workbook(f) for f in filepaths]
merged = Workbook()
o = merged.create_sheet('Sheet 4')
safeTitle = o.title
copySheet(filepath[0].worksheets[0],
          filepath[1].worksheets[0], merged[safeTitle])

merged.save('consolidated.xlsx')

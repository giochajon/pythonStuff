from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("test1gio")
ws.title = "Titulin"
ws3 = wb["Titulin"]
print (wb.sheetnames)
['Titulin', 'test1gio']
ws['A4']= 'prueba en 4'
wb.save('p1.xlsx')